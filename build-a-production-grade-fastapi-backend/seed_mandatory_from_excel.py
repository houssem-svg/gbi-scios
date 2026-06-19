"""Seed the mandatory_list_items table from the official Saudi government Excel.

Usage:
    python seed_mandatory_from_excel.py /path/to/القائمة الإلزامية.xlsx

If no path is given, defaults to /tmp/mandatory_list.xlsx.

The script:
  1. Opens the Excel workbook (19 sector sheets).
  2. For each sector sheet (skipping the "overview" and "exceptions" sheets):
     - Reads rows starting from row 2 (row 1 = headers).
     - Extracts: Etimad code, Arabic name, English name.
     - Uses the SHEET NAME as the sector.
  3. Cleans the data: strips whitespace, skips rows with no product name,
     skips header repetition rows.
  4. Clears the existing mandatory_list_items table (idempotent re-run).
  5. Bulk-inserts all products with penalty_percentage = 0.25 (25%).

Arabic text handling:
  - openpyxl reads cell values as Python str (Unicode) — no encoding issues.
  - We strip trailing whitespace and normalize Arabic letters (e.g. أ/إ/آ → ا)
    is NOT done here (kept for the fuzzy matcher to handle at query time).
"""
from __future__ import annotations

import sys
from decimal import Decimal
from pathlib import Path

# Ensure the backend package is importable when run as a script.
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl  # noqa: E402

from app.core.database import SessionLocal  # noqa: E402
from app.models.mandatory_list import (  # noqa: E402
    MandatoryListItem,
    MandatoryStatus,
)
from sqlalchemy import delete  # noqa: E402

# Sheets to SKIP (they are not product lists).
SKIP_SHEETS = {
    "نظرة عامة ",
    "نظرة عامة",
    "ضوابط الاستثناء العامة",
    "ضوابط الاستثناء لمنتجات الادوية",
}

# Column indices (1-based in Excel, 0-based in openpyxl values_only tuple).
# Verified against the official workbook header row.
COL_ETIMAD_CODE = 3   # col D (index 3) — "الرمز في منصة اعتماد"
COL_NAME_AR = 4       # col E (index 4) — "اسم المنتج (عربي)"
COL_NAME_EN = 5       # col F (index 5) — "اسم المنتج (إنجليزي)"
COL_SEGMENT_AR = 1    # col B (index 1) — "اسم القطاع الرئيسي (عربي)" (sub-category)


def _clean(value) -> str:
    """Strip whitespace and return empty string for None/NaN."""
    if value is None:
        return ""
    s = str(value).strip()
    return s


def _is_header_row(row: tuple) -> bool:
    """Detect a repeated header row inside the data (some sheets have them)."""
    name_ar = _clean(row[COL_NAME_AR]) if len(row) > COL_NAME_AR else ""
    return "اسم المنتج" in name_ar or "Commidity Title" in name_ar or "Commodity Title" in name_ar


def parse_workbook(xlsx_path: str) -> list[dict]:
    """Read all sector sheets and return a list of product dicts."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    products: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS or "ضوابط" in sheet_name or "نظرة" in sheet_name:
            continue

        ws = wb[sheet_name]
        sector = sheet_name.strip()
        current_segment = ""

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= COL_NAME_AR:
                continue

            # The segment title (col B) is only filled on the first row of each
            # segment group; subsequent rows leave it blank. Track it.
            segment_val = _clean(row[COL_SEGMENT_AR]) if len(row) > COL_SEGMENT_AR else ""
            if segment_val:
                current_segment = segment_val

            etimad_code = _clean(row[COL_ETIMAD_CODE]) if len(row) > COL_ETIMAD_CODE else ""
            name_ar = _clean(row[COL_NAME_AR]) if len(row) > COL_NAME_AR else ""
            name_en = _clean(row[COL_NAME_EN]) if len(row) > COL_NAME_EN else ""

            # Skip rows with no Arabic name (empty/separator/header rows).
            if not name_ar:
                continue
            # Skip repeated header rows.
            if _is_header_row(row):
                continue
            # Skip rows where the "name" is clearly not a product (e.g. notes).
            if len(name_ar) < 2:
                continue

            products.append(
                {
                    "item_code": etimad_code or f"{sector[:20]}-{len(products):04d}",
                    "product_name": name_ar,
                    "product_name_en": name_en or None,
                    "sector": sector,
                    "category": current_segment or sector,
                    "local_manufacturer": "",
                    "mandatory_status": MandatoryStatus.MANDATORY,
                    "penalty_percentage": Decimal("0.25"),
                }
            )

    return products


def main() -> None:
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "/tmp/mandatory_list.xlsx"
    if not Path(xlsx_path).exists():
        print(f"✗ File not found: {xlsx_path}")
        sys.exit(1)

    print(f"→ Reading workbook: {xlsx_path}")
    products = parse_workbook(xlsx_path)
    print(f"✓ Parsed {len(products)} products from sector sheets.")

    # Summary by sector
    by_sector: dict[str, int] = {}
    for p in products:
        by_sector[p["sector"]] = by_sector.get(p["sector"], 0) + 1
    for sector, count in sorted(by_sector.items(), key=lambda x: -x[1]):
        print(f"    {sector}: {count}")

    db = SessionLocal()
    try:
        # Clear existing items (idempotent re-run).
        db.execute(delete(MandatoryListItem))
        db.commit()
        print(f"✓ Cleared existing mandatory_list_items.")

        # Bulk insert in batches of 200.
        batch_size = 200
        inserted = 0
        for i in range(0, len(products), batch_size):
            batch = products[i : i + batch_size]
            db.bulk_save_objects([MandatoryListItem(**p) for p in batch])
            db.commit()
            inserted += len(batch)

        print(f"✓ Inserted {inserted} mandatory items.")
    finally:
        db.close()

    print("\nDone. Run a compliance scan to see fuzzy-matching in action.")


if __name__ == "__main__":
    main()
